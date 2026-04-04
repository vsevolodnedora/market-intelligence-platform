#!/usr/bin/env python3
from __future__ import annotations

import argparse
import asyncio
import base64
import io
import json
import re
import shutil
import sys
import traceback
import unicodedata
import zipfile
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from email.message import Message
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import unquote, urlparse

import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ── Constants ────────────────────────────────────────────────────────────────

DEFAULT_URL = "https://www.invesco.com/de/de/financial-products/etfs/invesco-ftse-all-world-ucits-etf-acc.html"
DEFAULT_ROLE = "Privatanleger"
DEFAULT_SECTION_TITLE = "Die 10 größten Positionen"
DEFAULT_EXPORT_TEXT = "Daten exportieren"
DEFAULT_ORIGINAL_FILENAME = "Die_10_größten_Positionen-holdings.xlsx"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
JSON_MIME = "application/json"
DEFAULT_OUTPUT_DIR = Path("../data/scrape_invesco-ftse-all-world/")


# ── Dataclasses / Exceptions ────────────────────────────────────────────────

@dataclass(slots=True)
class ScrapeManifest:
    source_url: str
    final_url: str | None
    as_of_date: str | None
    role_selected: str | None
    section_title: str
    export_text: str
    original_filename: str | None
    saved_file: str | None
    xlsx_candidate_urls: list[str]
    status_code: int | None
    success: bool
    partial_success: bool
    completeness_ok: bool | None
    expected_min_holdings: int | None
    trace_dir: str | None
    debug_files: list[str]
    holdings_summary: dict[str, Any] | None
    workbook_summary: dict[str, Any] | None
    notes: list[str]


@dataclass(slots=True)
class CapturedHoldings:
    """Metadata wrapper for a captured holdings JSON payload."""
    payload: dict[str, Any]
    url: str
    summary: dict[str, Any]
    post_click: bool
    holdings_count: int


class ScrapeError(RuntimeError):
    """Raised when the page could not be interacted with reliably."""


# ── Argument parsing ────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Traceable Invesco holdings scraper. Uses Crawl4AI + Playwright hooks "
            "to pass the Invesco investor gate, open the 'Positionen' tab, click "
            "the visible 'Daten exportieren' control for 'Die 10 größten Positionen', "
            "and save a valid holdings XLSX file. Includes screenshots, payload dumps, "
            "network logs, workbook inspection, and a minimum-holdings completeness check."
        )
    )
    parser.add_argument("--url", default=DEFAULT_URL)
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help=(
            "Root output directory. A date-stamped sub-directory YYYY_MM_DD/ "
            "will be created inside for each scrape run (replaced if it already "
            f"exists). Default: {DEFAULT_OUTPUT_DIR}"
        ),
    )
    parser.add_argument("--role", default=DEFAULT_ROLE)
    parser.add_argument("--section-title", default=DEFAULT_SECTION_TITLE)
    parser.add_argument("--export-text", default=DEFAULT_EXPORT_TEXT)
    parser.add_argument("--timeout-ms", type=int, default=90000)
    parser.add_argument(
        "--min-holdings",
        type=int,
        default=2500,
        help="Mark the scrape incomplete when fewer holdings than this are captured.",
    )
    parser.add_argument(
        "--trace-dir",
        default=None,
        help=(
            "Directory for screenshots, HTML/text snapshots, payload dumps, "
            "network logs, and the scrape log. Defaults to <run-output-dir>/trace."
        ),
    )
    parser.add_argument("--no-trace", action="store_true", help="Disable trace artifact capture.")
    parser.add_argument("--headed", action="store_true", help="Run a visible browser for debugging.")
    parser.add_argument("--verbose", action="store_true")
    parser.add_argument("--self-test", action="store_true", help="Run local validation tests only.")
    return parser.parse_args()


def resolve_run_output_dir(root_output_dir: Path) -> Path:
    """Create and return a date-stamped sub-directory ``YYYY_MM_DD/`` inside
    *root_output_dir*.  If the directory already exists it is removed first so
    the run starts clean."""
    today = datetime.now().strftime("%Y_%m_%d")
    run_dir = root_output_dir.resolve() / today
    if run_dir.exists():
        shutil.rmtree(run_dir)
    run_dir.mkdir(parents=True, exist_ok=True)
    return run_dir


# ── Utility helpers ─────────────────────────────────────────────────────────

def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def sanitize_filename(name: str) -> str:
    replacements = {
        "ä": "ae", "ö": "oe", "ü": "ue",
        "Ä": "Ae", "Ö": "Oe", "Ü": "Ue",
        "ß": "ss",
    }
    for src, dst in replacements.items():
        name = name.replace(src, dst)
    normalized = unicodedata.normalize("NFKD", name)
    normalized = normalized.encode("ascii", "ignore").decode("ascii")
    normalized = re.sub(r"[^A-Za-z0-9._-]+", "_", normalized)
    normalized = re.sub(r"_+", "_", normalized).strip("_")
    return normalized or "download.xlsx"


def extract_as_of_date_from_text(text: str, section_title: str = DEFAULT_SECTION_TITLE) -> str | None:
    patterns = [
        rf"{re.escape(section_title)}.*?Stand:\s*(\d{{2}}\.\d{{2}}\.\d{{4}})",
        r"Stand:\s*(\d{2}\.\d{2}\.\d{4})",
        r"As of:\s*(\d{2}\.\d{2}\.\d{4})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE | re.DOTALL)
        if match:
            return match.group(1)
    return None


def normalize_date_ddmmyyyy(date_str: str | None) -> str | None:
    if not date_str:
        return None
    date_str = date_str.strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%d.%m.%Y")
        except ValueError:
            continue
    return None


def choose_saved_filename(as_of_date: str | None, original_filename: str | None) -> str:
    original = original_filename or DEFAULT_ORIGINAL_FILENAME
    original = sanitize_filename(original)
    if as_of_date:
        yyyy, mm, dd = as_of_date[6:10], as_of_date[3:5], as_of_date[0:2]
        return f"{yyyy}-{mm}-{dd}__{original}"
    return original


def parse_content_disposition_filename(header_value: str | None) -> str | None:
    if not header_value:
        return None
    msg = Message()
    msg["content-disposition"] = header_value
    params = dict(msg.get_params(header="content-disposition", unquote=True) or [])
    filename = params.get("filename*") or params.get("filename")
    if not filename:
        return None
    if filename.lower().startswith("utf-8''"):
        filename = filename[7:]
    return Path(filename).name or None


def parse_filename_from_url(url: str | None) -> str | None:
    if not url:
        return None
    path = unquote(urlparse(url).path)
    name = Path(path).name
    return name or None


def extract_http_urls(text: str) -> list[str]:
    return re.findall(r"https?://[^\s\"'<>]+", text or "")


# ── XLSX / binary detection ─────────────────────────────────────────────────

def is_probable_xlsx_bytes(body: bytes | None) -> bool:
    if not body or len(body) < 4 or not body.startswith(b"PK"):
        return False
    try:
        with zipfile.ZipFile(io.BytesIO(body)) as archive:
            names = {name.lower() for name in archive.namelist()}
    except zipfile.BadZipFile:
        return False
    required = {"[content_types].xml", "_rels/.rels", "xl/workbook.xml"}
    return required.issubset(names)


def looks_like_holdings_json_url(url: str) -> bool:
    """Detect any holdings API URL, not just loadType=initial."""
    lower = url.lower()
    return "/holdings/" in lower


def _is_initial_load_url(url: str) -> bool:
    """Return True if the URL is a loadType=initial request (pre-click widget feed)."""
    return "loadtype=initial" in url.lower()


def _derive_full_holdings_urls(initial_url: str) -> list[str]:
    """Given a ``loadType=initial`` holdings URL, return candidate URLs for the
    full dataset — ``loadType=full`` and the URL with ``loadType`` removed.

    These are tried in order during the API-fetch fallback to obtain the
    complete holdings list rather than the truncated initial-widget feed.
    """
    candidates: list[str] = []
    if "loadtype=initial" in initial_url.lower():
        full_url = re.sub(r"loadType=initial", "loadType=full", initial_url, flags=re.I)
        candidates.append(full_url)
        # Also try removing loadType entirely — some API versions ignore unknown values
        no_load_type = re.sub(r"[&?]loadType=[^&]*", "", initial_url)
        if no_load_type != initial_url:
            candidates.append(no_load_type)
    return candidates


def looks_like_xlsx_response(
    *,
    url: str,
    content_type: str,
    content_disposition: str | None,
    body: bytes | None,
) -> bool:
    lower_url = url.lower()
    lower_type = (content_type or "").lower()
    filename = parse_content_disposition_filename(content_disposition) or parse_filename_from_url(url)
    if filename and filename.lower().endswith(".xlsx"):
        return True
    if XLSX_MIME in lower_type or "spreadsheetml" in lower_type:
        return True
    if lower_url.endswith(".xlsx"):
        return True
    return is_probable_xlsx_bytes(body)


def discover_candidate_urls(network_requests: Iterable[dict[str, Any]] | None) -> list[str]:
    if not network_requests:
        return []
    urls: list[str] = []
    for item in network_requests:
        raw_url = str(item.get("url") or "")
        if not raw_url:
            continue
        extracted = extract_http_urls(raw_url) or (
            [raw_url] if raw_url.startswith(("http://", "https://")) else []
        )
        headers = {str(k).lower(): str(v) for k, v in (item.get("headers") or {}).items()}
        content_type = headers.get("content-type", "").lower()
        for url in extracted:
            lower = url.lower()
            if (
                "/holdings/" in lower
                or lower.endswith(".xlsx")
                or XLSX_MIME in content_type
                or "spreadsheetml" in content_type
            ):
                urls.append(url)
    deduped: list[str] = []
    seen: set[str] = set()
    for url in urls:
        if url not in seen:
            seen.add(url)
            deduped.append(url)
    return deduped


def choose_holdings_json_url(urls: Iterable[str]) -> str | None:
    """Choose the best holdings JSON URL.

    Ranking (highest priority first):
      1. Post-click export/full holdings URLs (non-initial)
      2. Any holdings URL that is NOT loadType=initial
      3. loadType=initial (lowest priority, only as last resort)
    """
    candidates = [u for u in urls if "/holdings/" in u.lower()]
    if not candidates:
        return None

    def score(url: str) -> tuple[int, int, int, int]:
        lower = url.lower()
        # Deprioritize loadType=initial — it is the pre-click widget feed
        is_initial = 1 if "loadtype=initial" in lower else 0
        return (
            0 if is_initial else 1,  # non-initial first
            1 if "export" in lower or "download" in lower or "full" in lower else 0,
            1 if "idtype=isin" in lower else 0,
            1 if "accounts/de_de/shareclasses/" in lower else 0,
        )

    return sorted(candidates, key=score, reverse=True)[0]


# ── Holdings helpers ────────────────────────────────────────────────────────

def payload_is_authoritative(summary: dict[str, Any], min_holdings: int) -> bool:
    """Return True only if the payload passes minimum completeness checks.

    A payload is authoritative when it contains at least *min_holdings* rows
    and its declared total (if present) does not exceed the actual row count.
    """
    count = summary.get("holdings_count")
    if not isinstance(count, int):
        return False
    if count < min_holdings:
        return False
    # If the API declares more rows than it returned, the payload is partial
    declared_total = summary.get("declared_total")
    if isinstance(declared_total, int) and declared_total > count:
        return False
    return True


def rows_from_holdings_payload(payload: dict[str, Any]) -> list[dict[str, Any]]:
    holdings = payload.get("holdings")
    if not isinstance(holdings, list):
        raise ScrapeError("Holdings payload did not contain a 'holdings' list.")
    rows: list[dict[str, Any]] = []
    for item in holdings:
        if not isinstance(item, dict):
            continue
        rows.append(
            {
                "name": item.get("name"),
                "cusip": item.get("cusip"),
                "isin": item.get("isin"),
                "weight": item.get("weight"),
            }
        )
    if not rows:
        raise ScrapeError("Holdings payload was empty.")
    return rows


def summarize_holdings_payload(payload: dict[str, Any]) -> dict[str, Any]:
    holdings_raw = payload.get("holdings")
    holdings = [item for item in holdings_raw if isinstance(item, dict)] if isinstance(holdings_raw, list) else []
    weights = [float(item["weight"]) for item in holdings if isinstance(item.get("weight"), (int, float))]
    isins = [str(item.get("isin") or "").strip() for item in holdings if str(item.get("isin") or "").strip()]
    pagination_hints: dict[str, Any] = {}
    for key in ("totalCount", "totalRecords", "count", "page", "pageSize", "offset", "limit", "hasMore", "nextPage", "nextToken"):
        if key in payload:
            pagination_hints[key] = payload.get(key)
    for container_key in ("paging", "pagination", "meta"):
        container = payload.get(container_key)
        if isinstance(container, dict):
            interesting = {
                key: container.get(key)
                for key in ("totalCount", "totalRecords", "count", "page", "pageSize", "offset", "limit", "hasMore", "nextPage", "nextToken")
                if key in container
            }
            if interesting:
                pagination_hints[container_key] = interesting

    declared_total = None
    for key in ("totalCount", "totalRecords", "count"):
        value = payload.get(key)
        if isinstance(value, int):
            declared_total = value
            break
    if declared_total is None:
        for container_key in ("paging", "pagination", "meta"):
            container = payload.get(container_key)
            if isinstance(container, dict):
                for key in ("totalCount", "totalRecords", "count"):
                    value = container.get(key)
                    if isinstance(value, int):
                        declared_total = value
                        break
            if declared_total is not None:
                break

    return {
        "effective_date": normalize_date_ddmmyyyy(str(payload.get("effectiveDate") or "")),
        "holdings_count": len(holdings),
        "unique_isin_count": len(set(isins)),
        "weights_count": len(weights),
        "weight_sum_percent": round(sum(weights), 8) if weights else None,
        "min_weight_percent": round(min(weights), 8) if weights else None,
        "max_weight_percent": round(max(weights), 8) if weights else None,
        "payload_keys": sorted(str(key) for key in payload.keys()),
        "pagination_hints": pagination_hints or None,
        "declared_total": declared_total,
        "declared_total_exceeds_rows": bool(declared_total and declared_total > len(holdings)),
        "sample_names": [item.get("name") for item in holdings[:5]],
    }


def inspect_workbook(path: str | Path | None) -> dict[str, Any] | None:
    if not path:
        return None
    workbook_path = Path(path)
    if not workbook_path.exists():
        return None
    summary: dict[str, Any] = {
        "path": str(workbook_path.resolve()),
        "file_size_bytes": workbook_path.stat().st_size,
    }
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    summary["sheet_names"] = list(wb.sheetnames)
    if "Holdings" in wb.sheetnames:
        ws = wb["Holdings"]
        row_count = 0
        sample_names: list[str] = []
        weights: list[float] = []
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row or not any(value not in (None, "") for value in row[:4]):
                continue
            row_count += 1
            if len(sample_names) < 5:
                sample_names.append(str(row[0]))
            weight = row[3] if len(row) > 3 else None
            if isinstance(weight, (int, float)):
                weights.append(float(weight) * 100.0 if float(weight) <= 1.0 else float(weight))
        summary.update(
            {
                "data_row_count": row_count,
                "weight_sum_percent": round(sum(weights), 8) if weights else None,
                "sample_names": sample_names,
            }
        )
    if "Metadata" in wb.sheetnames:
        meta_sheet = wb["Metadata"]
        metadata_map: dict[str, Any] = {}
        for row in meta_sheet.iter_rows(min_row=1, max_col=2, values_only=True):
            key, value = row[:2]
            if key:
                metadata_map[str(key)] = value
        summary["metadata"] = metadata_map
    return summary


def evaluate_completeness(
    *,
    workbook_summary: dict[str, Any] | None,
    holdings_summary: dict[str, Any] | None,
    min_holdings: int,
) -> tuple[bool | None, list[str]]:
    reasons: list[str] = []
    workbook_rows = workbook_summary.get("data_row_count") if workbook_summary else None
    if isinstance(workbook_rows, int) and workbook_rows < min_holdings:
        reasons.append(f"workbook contains {workbook_rows} holdings, below the minimum expected {min_holdings}")
    payload_rows = holdings_summary.get("holdings_count") if holdings_summary else None
    if isinstance(payload_rows, int) and payload_rows < min_holdings:
        reasons.append(f"API payload contains {payload_rows} holdings, below the minimum expected {min_holdings}")
    declared_total = holdings_summary.get("declared_total") if holdings_summary else None
    if isinstance(payload_rows, int) and isinstance(declared_total, int) and declared_total > payload_rows:
        reasons.append(f"API payload declares {declared_total} total holdings but only returned {payload_rows} rows")
    if reasons:
        return False, reasons
    if workbook_summary or holdings_summary:
        return True, []
    return None, []


# ── Trace logger ────────────────────────────────────────────────────────────

class TraceLogger:
    def __init__(self, trace_dir: Path | None, *, verbose: bool = False) -> None:
        self.trace_dir = ensure_dir(trace_dir.resolve()) if trace_dir else None
        self.verbose = verbose
        self.files: list[str] = []
        self.log_path = self.trace_dir / "scrape.log" if self.trace_dir else None
        if self.log_path:
            self.log_path.write_text("", encoding="utf-8")
            self.remember(self.log_path)

    def log(self, message: str) -> None:
        line = f"[{utc_now_iso()}] {message}"
        if self.log_path:
            with self.log_path.open("a", encoding="utf-8") as handle:
                handle.write(line + "\n")
        if self.verbose:
            print(line, file=sys.stderr)

    def remember(self, path: str | Path | None) -> str | None:
        if not path:
            return None
        resolved = str(Path(path).resolve())
        if resolved not in self.files:
            self.files.append(resolved)
        return resolved

    def save_text(self, relative_name: str, content: str) -> str | None:
        if not self.trace_dir:
            return None
        path = self.trace_dir / relative_name
        ensure_dir(path.parent)
        path.write_text(content, encoding="utf-8")
        return self.remember(path)

    def save_json(self, relative_name: str, payload: Any) -> str | None:
        if not self.trace_dir:
            return None
        path = self.trace_dir / relative_name
        ensure_dir(path.parent)
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        return self.remember(path)


# ── Playwright interaction helpers ──────────────────────────────────────────

async def wait_for_first_visible(candidates: Iterable[Any], timeout_ms: int) -> Any | None:
    for locator in candidates:
        try:
            await locator.first.wait_for(state="visible", timeout=timeout_ms)
            return locator.first
        except Exception:
            continue
    return None


async def click_if_present(candidates: Iterable[Any], timeout_ms: int, force: bool = False) -> bool:
    locator = await wait_for_first_visible(candidates, timeout_ms)
    if locator is None:
        return False
    return await click_export_locator(locator, timeout_ms=timeout_ms, force=force)


async def click_export_locator(locator: Any, timeout_ms: int, force: bool = True) -> bool:
    try:
        await locator.scroll_into_view_if_needed(timeout=timeout_ms)
    except Exception:
        pass

    try:
        await locator.click(timeout=timeout_ms, force=force)
        return True
    except Exception:
        pass

    try:
        await locator.evaluate(
            r'''
            (el) => {
                const clickable = el.closest('button, a, [role="button"], [download], [tabindex], .download__button') || el;
                const fire = (type, ctor = MouseEvent) => clickable.dispatchEvent(new ctor(type, {
                    bubbles: true,
                    cancelable: true,
                    composed: true,
                    view: window,
                }));
                fire('pointerdown', PointerEvent);
                fire('mousedown');
                fire('pointerup', PointerEvent);
                fire('mouseup');
                fire('click');
                return true;
            }
            '''
        )
        return True
    except Exception:
        return False


async def maybe_accept_cookie_banner(page: Any, timeout_ms: int) -> None:
    candidates = [
        page.get_by_role("button", name=re.compile(r"^(Alle akzeptieren|Akzeptieren|Accept all|Accept)$", re.I)),
        page.get_by_role("link", name=re.compile(r"^(Alle akzeptieren|Akzeptieren|Accept all|Accept)$", re.I)),
        page.get_by_text(re.compile(r"^(Alle akzeptieren|Akzeptieren|Accept all|Accept)$", re.I)),
    ]
    try:
        await click_if_present(candidates, min(timeout_ms, 4000))
    except Exception:
        pass


async def handle_investor_gate(page: Any, role_text: str, timeout_ms: int, notes: list[str]) -> None:
    """Detect the Invesco investor-type gate, select a role, and confirm.

    The gate is a React overlay that asks "Welcher Anlegertyp sind Sie?" and
    requires both *selecting* a role (radio / button) and *confirming* it via a
    "Bestätigen" button.  The confirmation button may only become enabled or
    visible after the role is selected, so we wait briefly after the click.
    """
    gate_markers = [
        page.get_by_text(re.compile(r"Welcher Anlegertyp sind Sie\?", re.I)),
        page.get_by_text(re.compile(r"Bitte bestätigen Sie ihren Anlegertyp", re.I)),
        page.get_by_text(re.compile(r"Bitte wählen Sie aus, zu welcher Anlegergruppe", re.I)),
        # CSS fallback — the splash overlay typically lives in a recognisable container
        page.locator(".country-splash, .splash-selector, [class*='splash'], [class*='gate'], [class*='role-selector']"),
    ]
    gate = await wait_for_first_visible(gate_markers, min(timeout_ms, 8000))
    if gate is None:
        notes.append("Investor gate did not appear; continuing without role selection.")
        return

    # ── Select the investor role ──
    role_candidates = [
        page.get_by_role("radio", name=re.compile(fr"^{re.escape(role_text)}$", re.I)),
        page.get_by_role("button", name=re.compile(fr"^{re.escape(role_text)}$", re.I)),
        page.get_by_role("link", name=re.compile(fr"^{re.escape(role_text)}$", re.I)),
        page.get_by_label(re.compile(fr"^{re.escape(role_text)}$", re.I)),
        page.get_by_text(re.compile(fr"^{re.escape(role_text)}$", re.I)),
        # Partial / substring match as fallback
        page.locator(f"label:has-text('{role_text}'), [class*='radio']:has-text('{role_text}')"),
    ]
    if not await click_if_present(role_candidates, min(timeout_ms, 12000), force=True):
        raise ScrapeError(f"Could not select investor role: {role_text}")
    notes.append(f"Selected investor role: {role_text}")

    # Give the page time to react (the confirm button may appear or enable after selection)
    await page.wait_for_timeout(1500)

    # ── Click the confirm button ──
    # The button text on the German site is "Bestätigen" but may also read
    # "Confirm", "Website besuchen", or be inside a submit-type element.
    confirm_candidates = [
        page.get_by_role("button", name=re.compile(r"^(Bestätigen|Confirm|Website besuchen)$", re.I)),
        page.get_by_role("link", name=re.compile(r"^(Bestätigen|Confirm|Website besuchen)$", re.I)),
        # Substring match — the button might contain extra whitespace or an icon
        page.get_by_role("button", name=re.compile(r"Bestätigen|Confirm|Website besuchen", re.I)),
        page.get_by_role("link", name=re.compile(r"Bestätigen|Confirm|Website besuchen", re.I)),
        page.get_by_text(re.compile(r"^(Bestätigen|Confirm|Website besuchen)$", re.I)),
        # CSS-based selectors for common Invesco splash patterns
        page.locator("button:has-text('Bestätigen'), a:has-text('Bestätigen')"),
        page.locator("button:has-text('Confirm'), a:has-text('Confirm')"),
        page.locator("button[type='submit'], input[type='submit']"),
        page.locator(".splash-confirm, .country-splash button, [class*='splash'] button"),
    ]
    confirmed = await click_if_present(confirm_candidates, min(timeout_ms, 12000), force=True)

    if not confirmed:
        # Last resort: use JS to find and click the most likely confirm element
        try:
            confirmed = await page.evaluate(
                r'''
                () => {
                    const candidates = [...document.querySelectorAll('button, a, [role="button"], input[type="submit"]')];
                    const confirm = candidates.find(el => {
                        const text = (el.innerText || el.textContent || '').trim().toLowerCase();
                        return /bestätigen|confirm|website besuchen/.test(text);
                    });
                    if (confirm) { confirm.click(); return true; }
                    return false;
                }
                '''
            )
        except Exception:
            confirmed = False

    if confirmed:
        notes.append("Confirmed investor role via gate button.")
    else:
        notes.append("Role selected, but no explicit confirm button was found.")

    # Wait for the page to settle after confirmation
    for load_state in ("domcontentloaded", "load", "networkidle"):
        try:
            await page.wait_for_load_state(load_state, timeout=min(timeout_ms, 10000))
        except Exception:
            pass
    await page.wait_for_timeout(2000)


async def open_positionen_tab_if_needed(page: Any, section_title: str, timeout_ms: int, notes: list[str]) -> None:
    title_locators = [
        page.get_by_role("heading", name=re.compile(fr"^{re.escape(section_title)}$", re.I)),
        page.get_by_text(re.compile(fr"^{re.escape(section_title)}$", re.I)),
    ]
    if await wait_for_first_visible(title_locators, 3000):
        return

    tab_candidates = [
        page.get_by_role("tab", name=re.compile(r"^Positionen$", re.I)),
        page.get_by_role("button", name=re.compile(r"^Positionen$", re.I)),
        page.get_by_role("link", name=re.compile(r"^Positionen$", re.I)),
        page.get_by_text(re.compile(r"^Positionen$", re.I)),
    ]
    clicked = await click_if_present(tab_candidates, min(timeout_ms, 10000), force=True)
    if clicked:
        notes.append("Opened the 'Positionen' tab/section.")
        await page.wait_for_timeout(1500)


async def wait_for_holdings_section(page: Any, section_title: str, timeout_ms: int) -> None:
    title_pattern = re.escape(section_title)
    js = (
        "() => {"
        " const text = document.body ? document.body.innerText : '';"
        rf" return /{title_pattern}/i.test(text) && /(Stand:|As of:)\s*\d{{2}}\.\d{{2}}\.\d{{4}}/i.test(text);"
        " }"
    )
    await page.wait_for_function(js, timeout=timeout_ms)


async def extract_as_of_date(page: Any, section_title: str) -> str | None:
    try:
        extracted = await page.evaluate(
            r'''
            (sectionTitle) => {
                const bodyText = document.body ? document.body.innerText : "";
                const headingCandidates = [...document.querySelectorAll('h1,h2,h3,h4,h5,h6,div,span,p')]
                    .filter(el => (el.textContent || '').trim() === sectionTitle);
                for (const heading of headingCandidates) {
                    let node = heading;
                    for (let i = 0; i < 8 && node; i += 1) {
                        const text = node.innerText || '';
                        const m = text.match(/Stand:\s*(\d{2}\.\d{2}\.\d{4})/i) || text.match(/As of:\s*(\d{2}\.\d{2}\.\d{4})/i);
                        if (m) return m[1];
                        node = node.parentElement;
                    }
                }
                const globalMatch = bodyText.match(/Stand:\s*(\d{2}\.\d{2}\.\d{4})/i) || bodyText.match(/As of:\s*(\d{2}\.\d{2}\.\d{4})/i);
                return globalMatch ? globalMatch[1] : null;
            }
            ''',
            section_title,
        )
        normalized = normalize_date_ddmmyyyy(str(extracted)) if extracted else None
        if normalized:
            return normalized
    except Exception:
        pass
    html = await page.content()
    return extract_as_of_date_from_text(html, section_title=section_title)


async def find_section_export_locator(page: Any, section_title: str, export_text: str) -> Any | None:
    section_xpath = (
        "xpath=(//*[self::h1 or self::h2 or self::h3 or self::h4 or self::h5 or self::h6 or self::div or self::span or self::p]"
        f"[normalize-space()={json.dumps(section_title)}]"
        "/ancestor::*[.//*[contains(normalize-space(.),"
        f" {json.dumps(export_text)} )]][1]"
        "//*[self::button or self::a or @role='button' or contains(@class,'download__button') or contains(@class,'download__button__text')]"
        f"[contains(normalize-space(.), {json.dumps(export_text)})])[1]"
    )
    nearby_xpath = (
        "xpath=(//*[self::h1 or self::h2 or self::h3 or self::h4 or self::h5 or self::h6 or self::div or self::span or self::p]"
        f"[normalize-space()={json.dumps(section_title)}]"
        "/following::*[self::button or self::a or @role='button' or contains(@class,'download__button') or contains(@class,'download__button__text')]"
        f"[contains(normalize-space(.), {json.dumps(export_text)})][1])"
    )
    candidates = [
        page.locator(section_xpath),
        page.locator(nearby_xpath),
        page.locator("button:has-text('Daten exportieren'), a:has-text('Daten exportieren')"),
        page.locator(".download__button, .download__button__text", has_text=re.compile(export_text, re.I)),
        page.get_by_role("button", name=re.compile(fr"^{re.escape(export_text)}$", re.I)),
        page.get_by_role("link", name=re.compile(fr"^{re.escape(export_text)}$", re.I)),
        page.get_by_text(re.compile(fr"^{re.escape(export_text)}$", re.I)),
    ]
    return await wait_for_first_visible(candidates, 8000)


async def trigger_export_download(page: Any, section_title: str, export_text: str, timeout_ms: int, notes: list[str]) -> Any | None:
    """Attempt to trigger the export/download action and return a Playwright
    Download object if a native browser download was captured.

    The Invesco product page often uses a React-based download component that
    builds the file client-side as a blob URL and triggers an anchor click
    rather than a direct server download.  When the native download path fails,
    this function still performs the click so that the blob-capture hooks and
    network interceptors can pick up the result.
    """
    locator = await find_section_export_locator(page, section_title, export_text)
    if locator is None:
        notes.append("Could not find export control locator — skipping export click.")
        return None

    # ── Attempt 1: native browser download via Playwright ──
    for click_mode in ("playwright", "dom-dispatch"):
        try:
            async with page.expect_download(timeout=min(timeout_ms, 12000)) as download_info:
                if click_mode == "playwright":
                    await click_export_locator(locator, timeout_ms=timeout_ms, force=True)
                else:
                    await locator.evaluate(
                        r'''
                        (el) => {
                            const clickable = el.closest('button, a, [role="button"], [download], [tabindex], .download__button') || el;
                            clickable.click();
                        }
                        '''
                    )
            notes.append(f"Triggered export control via {click_mode} click path (native download).")
            return await download_info.value
        except Exception:
            await page.wait_for_timeout(800)

    # ── Attempt 2: click without expecting a download ──
    # Many React SPAs generate the file client-side and trigger a blob download.
    # The click still needs to happen so the blob hooks capture it.
    notes.append(
        "No native browser download event — performing a plain click for "
        "blob/client-side download capture."
    )
    try:
        await click_export_locator(locator, timeout_ms=timeout_ms, force=True)
        # Give the React component time to fetch data and build the blob
        await page.wait_for_timeout(5000)
    except Exception as exc:
        notes.append(f"Plain click on export locator also failed: {exc}")

    return None


async def save_playwright_download(download: Any, dest_dir: Path, as_of_date: str | None) -> tuple[str, str]:
    original_filename = getattr(download, "suggested_filename", None) or DEFAULT_ORIGINAL_FILENAME
    saved_filename = choose_saved_filename(as_of_date, original_filename)
    dest_path = dest_dir / saved_filename
    await download.save_as(str(dest_path))
    body = dest_path.read_bytes()
    if not is_probable_xlsx_bytes(body):
        dest_path.unlink(missing_ok=True)
        raise ScrapeError(f"Downloaded file was not a valid XLSX workbook: {original_filename}")
    return original_filename, str(dest_path.resolve())


async def try_save_blob_download(page: Any, dest_dir: Path, as_of_date: str | None) -> tuple[str, str] | None:
    try:
        audit = await page.evaluate("window.__invescoDownloadAudit || { anchorClicks: [], blobUrls: [] }")
    except Exception:
        return None

    blob_urls: list[tuple[str, str | None]] = []
    for anchor in audit.get("anchorClicks") or []:
        href = str(anchor.get("href") or "")
        if href.startswith("blob:"):
            filename = str(anchor.get("download") or "") or None
            blob_urls.append((href, filename))
    for item in audit.get("blobUrls") or []:
        href = str(item.get("url") or "")
        if href.startswith("blob:"):
            blob_urls.append((href, None))

    seen: set[str] = set()
    for href, hinted_name in blob_urls:
        if href in seen:
            continue
        seen.add(href)
        try:
            encoded = await page.evaluate(
                r'''
                async (blobUrl) => {
                    const resp = await fetch(blobUrl);
                    const buffer = await resp.arrayBuffer();
                    const bytes = new Uint8Array(buffer);
                    let binary = '';
                    const chunk = 0x8000;
                    for (let i = 0; i < bytes.length; i += chunk) {
                        binary += String.fromCharCode(...bytes.subarray(i, i + chunk));
                    }
                    return btoa(binary);
                }
                ''',
                href,
            )
            if not encoded:
                continue
            body = base64.b64decode(encoded)
            if not is_probable_xlsx_bytes(body):
                continue
            original_filename = hinted_name or DEFAULT_ORIGINAL_FILENAME
            saved_filename = choose_saved_filename(as_of_date, original_filename)
            dest_path = dest_dir / saved_filename
            dest_path.write_bytes(body)
            return original_filename, str(dest_path.resolve())
        except Exception:
            continue
    return None


# ── HTTP fallback helpers ───────────────────────────────────────────────────

async def fallback_fetch_binary(
    url: str,
    cookies: list[dict[str, Any]],
    user_agent: str | None,
) -> tuple[bytes | None, str | None, str | None]:
    jar = httpx.Cookies()
    for cookie in cookies:
        name = cookie.get("name")
        value = cookie.get("value")
        domain = cookie.get("domain")
        path = cookie.get("path") or "/"
        if name and value and domain:
            jar.set(name, value, domain=domain, path=path)

    headers = {"Accept-Language": "de-DE,de;q=0.9,en;q=0.8"}
    if user_agent:
        headers["User-Agent"] = user_agent

    async with httpx.AsyncClient(follow_redirects=True, timeout=60, cookies=jar, headers=headers) as client:
        response = await client.get(url)
        response.raise_for_status()
        content_type = (response.headers.get("content-type") or "").lower()
        content_disposition = response.headers.get("content-disposition")
        filename = parse_content_disposition_filename(content_disposition) or parse_filename_from_url(str(response.url))
        body = response.content
        if looks_like_xlsx_response(
            url=str(response.url),
            content_type=content_type,
            content_disposition=content_disposition,
            body=body,
        ):
            return body, filename, content_type
        return None, filename, content_type


async def fetch_json_from_url(
    url: str,
    cookies: list[dict[str, Any]],
    user_agent: str | None,
) -> dict[str, Any]:
    jar = httpx.Cookies()
    for cookie in cookies:
        name = cookie.get("name")
        value = cookie.get("value")
        domain = cookie.get("domain")
        path = cookie.get("path") or "/"
        if name and value and domain:
            jar.set(name, value, domain=domain, path=path)

    headers = {
        "Accept-Language": "de-DE,de;q=0.9,en;q=0.8",
        "Accept": "application/json, text/plain, */*",
    }
    if user_agent:
        headers["User-Agent"] = user_agent

    async with httpx.AsyncClient(follow_redirects=True, timeout=60, cookies=jar, headers=headers) as client:
        response = await client.get(url)
        response.raise_for_status()
        return response.json()


# ── XLSX builder ────────────────────────────────────────────────────────────

def build_xlsx_from_holdings_payload(
    payload: dict[str, Any],
    source_url: str,
    holdings_url: str | None,
    section_title: str,
    output_path: Path,
) -> None:
    rows = rows_from_holdings_payload(payload)
    effective_date = normalize_date_ddmmyyyy(str(payload.get("effectiveDate") or ""))

    wb = Workbook()
    ws = wb.active
    ws.title = "Holdings"

    ws["A1"] = section_title
    ws["A2"] = f"Stand: {effective_date or ''}".strip()
    ws["A4"] = "Bezeichnung"
    ws["B4"] = "CUSIP"
    ws["C4"] = "ISIN"
    ws["D4"] = "Gewichtung"

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    start_row = 5
    for idx, item in enumerate(rows, start=start_row):
        ws[f"A{idx}"] = item.get("name")
        ws[f"B{idx}"] = item.get("cusip")
        ws[f"C{idx}"] = item.get("isin")
        weight = item.get("weight")
        if isinstance(weight, (int, float)):
            ws[f"D{idx}"] = weight / 100.0
            ws[f"D{idx}"].number_format = "0.0000%"
        else:
            ws[f"D{idx}"] = weight

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:D{ws.max_row}"
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14

    metadata = wb.create_sheet("Metadata")
    metadata["A1"] = "source_url"
    metadata["B1"] = source_url
    metadata["A2"] = "holdings_api_url"
    metadata["B2"] = holdings_url or ""
    metadata["A3"] = "effective_date"
    metadata["B3"] = effective_date or ""
    metadata["A4"] = "generated_at_utc"
    metadata["B4"] = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    metadata.column_dimensions["A"].width = 22
    metadata.column_dimensions["B"].width = 120

    wb.save(output_path)


# ── Trace snapshot helpers ──────────────────────────────────────────────────

async def save_page_snapshot(
    page: Any,
    trace: TraceLogger,
    label: str,
    *,
    include_html: bool = False,
    include_text: bool = False,
) -> dict[str, str]:
    artifacts: dict[str, str] = {}
    if not trace.trace_dir:
        return artifacts
    base_name = sanitize_filename(label)
    screenshot_path = trace.trace_dir / f"{base_name}.png"
    try:
        await page.screenshot(path=str(screenshot_path), full_page=True)
        saved = trace.remember(screenshot_path)
        if saved:
            artifacts["screenshot"] = saved
    except Exception as exc:
        trace.log(f"Failed to save screenshot for {label}: {exc}")
    if include_html:
        try:
            html = await page.content()
            saved = trace.save_text(f"{base_name}.html", html)
            if saved:
                artifacts["html"] = saved
        except Exception as exc:
            trace.log(f"Failed to save HTML snapshot for {label}: {exc}")
    if include_text:
        try:
            text = await page.evaluate("document.body ? document.body.innerText : ''")
            saved = trace.save_text(f"{base_name}.txt", text)
            if saved:
                artifacts["text"] = saved
        except Exception as exc:
            trace.log(f"Failed to save text snapshot for {label}: {exc}")
    return artifacts


async def collect_page_debug_summary(page: Any, section_title: str, export_text: str) -> dict[str, Any]:
    """Collect a structured page-state summary for diagnostics.

    Note: Playwright's ``page.evaluate(expression, arg)`` accepts at most ONE
    *arg* parameter.  We therefore pack both values into a list and destructure
    on the JS side.
    """
    return await page.evaluate(
        r'''
        ([sectionTitle, exportText]) => {
            const bodyText = document.body ? document.body.innerText : '';
            const exportNodes = [...document.querySelectorAll('a, button, [role="button"], span.download__button__text')]
                .filter(el => (el.innerText || el.textContent || '').toLowerCase().includes(exportText.toLowerCase()))
                .slice(0, 12)
                .map((el) => ({
                    tag: el.tagName,
                    text: (el.innerText || el.textContent || '').trim(),
                    href: el.href || '',
                    download: el.download || '',
                    className: el.className || '',
                    ariaLabel: el.getAttribute('aria-label') || '',
                }));
            const headings = [...document.querySelectorAll('h1,h2,h3,h4,h5,h6')]
                .map((el) => (el.innerText || '').trim())
                .filter(Boolean)
                .slice(0, 30);
            const activeTabs = [...document.querySelectorAll('[role="tab"], a, button')]
                .filter(el => /positionen/i.test((el.innerText || el.textContent || '').trim()))
                .slice(0, 12)
                .map(el => ({
                    text: (el.innerText || el.textContent || '').trim(),
                    ariaSelected: el.getAttribute('aria-selected'),
                    className: el.className || '',
                }));
            const sectionIndex = bodyText.indexOf(sectionTitle);
            const sectionSnippet = sectionIndex >= 0 ? bodyText.slice(Math.max(0, sectionIndex - 250), Math.min(bodyText.length, sectionIndex + 1500)) : null;
            return {
                title: document.title || '',
                url: window.location.href,
                hasSectionTitle: bodyText.includes(sectionTitle),
                hasExportText: bodyText.toLowerCase().includes(exportText.toLowerCase()),
                headings,
                activeTabs,
                exportNodes,
                sectionSnippet,
            };
        }
        ''',
        [section_title, export_text],
    )


# ── Captured holdings selection ─────────────────────────────────────────────

def choose_best_captured_holdings(
    captured_list: list[CapturedHoldings],
    min_holdings: int,
) -> CapturedHoldings | None:
    """Pick the best holdings payload from all captured network responses.

    Priority order:
      1. Post-click authoritative payload (meets min_holdings, no partial markers)
      2. Post-click payload (largest row count)
      3. Pre-click authoritative payload
      4. Pre-click payload (largest row count) — only if authoritative
      5. None (refuse to use non-authoritative pre-click payloads)
    """
    if not captured_list:
        return None

    # Separate post-click from pre-click
    post_click = [c for c in captured_list if c.post_click]
    pre_click = [c for c in captured_list if not c.post_click]

    # Among post-click: prefer authoritative, then largest
    post_click_auth = [c for c in post_click if payload_is_authoritative(c.summary, min_holdings)]
    if post_click_auth:
        return max(post_click_auth, key=lambda c: c.holdings_count)

    # Any post-click payload (even partial) is better than pre-click initial
    if post_click:
        return max(post_click, key=lambda c: c.holdings_count)

    # Pre-click: only use if authoritative (i.e. meets completeness)
    pre_click_auth = [c for c in pre_click if payload_is_authoritative(c.summary, min_holdings)]
    if pre_click_auth:
        return max(pre_click_auth, key=lambda c: c.holdings_count)

    # Refuse to use non-authoritative pre-click payloads (the root cause fix)
    return None


# ── Main scrape orchestrator ────────────────────────────────────────────────

async def run_scrape(args: argparse.Namespace) -> ScrapeManifest:
    from crawl4ai import AsyncWebCrawler, BrowserConfig, CacheMode, CrawlerRunConfig

    output_dir = resolve_run_output_dir(Path(args.output_dir))
    trace_dir = None if args.no_trace else Path(args.trace_dir).resolve() if args.trace_dir else output_dir / "trace"
    trace = TraceLogger(trace_dir, verbose=args.verbose)
    notes: list[str] = []

    def note(message: str) -> None:
        notes.append(message)
        trace.log(message)

    state: dict[str, Any] = {
        "as_of_date": None,
        "saved_path": None,
        "original_filename": None,
        "cookies": [],
        "user_agent": None,
        "final_url": None,
        "holdings_payload": None,
        "holdings_url": None,
        "holdings_summary": None,
        "captured_binary": None,
        "response_tasks": [],
        "network_events": [],
        "console_messages": [],
        "export_audit": None,
        "workbook_summary": None,
        # ── New state for multi-payload capture ──
        "captured_holdings": [],       # list[CapturedHoldings]
        "export_click_fired": False,   # set True after the export button is clicked
    }

    trace.log(f"Starting scrape for {args.url}")
    trace.log(f"Output directory: {output_dir}")
    if trace.trace_dir:
        trace.log(f"Trace directory: {trace.trace_dir}")

    browser_config = BrowserConfig(
        browser_type="chromium",
        headless=not args.headed,
        accept_downloads=True,
        downloads_path=str(output_dir),
        enable_stealth=True,
        user_agent_mode="random",
        viewport_width=1440,
        viewport_height=2200,
        headers={"Accept-Language": "de-DE,de;q=0.9,en;q=0.8"},
        verbose=args.verbose,
        extra_args=[
            "--disable-blink-features=AutomationControlled",
            "--lang=de-DE",
        ],
    )

    run_config = CrawlerRunConfig(
        cache_mode=CacheMode.BYPASS,
        wait_until="domcontentloaded",
        wait_for="css:body",
        wait_for_timeout=args.timeout_ms,
        page_timeout=args.timeout_ms,
        delay_before_return_html=1.5,
        scan_full_page=True,
        scroll_delay=0.3,
        flatten_shadow_dom=True,
        remove_consent_popups=True,
        simulate_user=True,
        override_navigator=True,
        capture_network_requests=True,
        capture_console_messages=True,
        verbose=args.verbose,
    )

    crawler = AsyncWebCrawler(config=browser_config)

    async def capture_response(response: Any) -> None:
        try:
            url = response.url
            status = response.status
            headers = {k.lower(): v for k, v in (await response.all_headers()).items()}
            content_type = headers.get("content-type", "")
            content_disposition = headers.get("content-disposition")
            event = {
                "event": "response",
                "timestamp_utc": utc_now_iso(),
                "status": status,
                "url": url,
                "content_type": content_type,
                "content_disposition": content_disposition,
            }
            state["network_events"].append(event)
            if len(state["network_events"]) <= 60 or "/holdings/" in url.lower() or ".xlsx" in url.lower():
                trace.log(f"HTTP {status} {url} [{content_type or 'unknown'}]")

            # ── Capture ALL holdings JSON payloads (not just the first) ──
            if looks_like_holdings_json_url(url):
                try:
                    payload = await response.json()
                    if isinstance(payload, dict) and payload.get("holdings"):
                        summary = summarize_holdings_payload(payload)
                        post_click = state["export_click_fired"]
                        captured = CapturedHoldings(
                            payload=payload,
                            url=url,
                            summary=summary,
                            post_click=post_click,
                            holdings_count=summary.get("holdings_count", 0),
                        )
                        state["captured_holdings"].append(captured)
                        trace.log(
                            f"Captured holdings payload ({'post-click' if post_click else 'pre-click'}): "
                            f"rows={summary.get('holdings_count')} "
                            f"initial={'yes' if _is_initial_load_url(url) else 'no'} "
                            f"url={url}"
                        )
                except Exception as exc:
                    trace.log(f"Failed to parse holdings JSON from {url}: {exc}")

            should_probe_binary = (
                ".xlsx" in url.lower()
                or XLSX_MIME in content_type.lower()
                or "spreadsheetml" in content_type.lower()
                or (content_disposition and "attachment" in content_disposition.lower())
                or "octet-stream" in content_type.lower()
            )
            if not should_probe_binary or state.get("captured_binary"):
                return

            body = await response.body()
            if not looks_like_xlsx_response(
                url=url,
                content_type=content_type,
                content_disposition=content_disposition,
                body=body,
            ):
                trace.log(f"Ignored non-XLSX binary candidate from {url}")
                return

            state["captured_binary"] = {
                "url": url,
                "body": body,
                "content_type": content_type,
                "filename": (
                    parse_content_disposition_filename(content_disposition)
                    or parse_filename_from_url(url)
                    or DEFAULT_ORIGINAL_FILENAME
                ),
            }
            trace.log(f"Captured validated XLSX-like network response from {url}")
        except Exception as exc:
            trace.log(f"capture_response error: {exc}")

    async def on_page_context_created(page: Any, context: Any, **_: Any) -> Any:
        context.set_default_timeout(args.timeout_ms)
        context.set_default_navigation_timeout(args.timeout_ms)
        await page.set_viewport_size({"width": 1440, "height": 2200})
        await context.add_init_script(
            r'''
            (() => {
                window.__invescoDownloadAudit = {
                    anchorClicks: [],
                    blobUrls: [],
                    windowOpens: []
                };

                const origCreateObjectURL = URL.createObjectURL.bind(URL);
                URL.createObjectURL = function(obj) {
                    const url = origCreateObjectURL(obj);
                    try {
                        window.__invescoDownloadAudit.blobUrls.push({
                            url,
                            type: obj && obj.type ? obj.type : '',
                            size: obj && typeof obj.size === 'number' ? obj.size : null
                        });
                    } catch (_) {}
                    return url;
                };

                const origAnchorClick = HTMLAnchorElement.prototype.click;
                HTMLAnchorElement.prototype.click = function(...args) {
                    try {
                        window.__invescoDownloadAudit.anchorClicks.push({
                            href: this.href || '',
                            download: this.download || '',
                            text: (this.textContent || '').trim(),
                            target: this.target || ''
                        });
                    } catch (_) {}
                    return origAnchorClick.apply(this, args);
                };

                const origOpen = window.open;
                window.open = function(...args) {
                    try {
                        window.__invescoDownloadAudit.windowOpens.push({
                            url: args[0] || '',
                            target: args[1] || ''
                        });
                    } catch (_) {}
                    return origOpen.apply(this, args);
                };
            })();
            '''
        )
        page.on("response", lambda response: state["response_tasks"].append(asyncio.create_task(capture_response(response))))
        page.on(
            "console",
            lambda message: state["console_messages"].append(
                {
                    "timestamp_utc": utc_now_iso(),
                    "type": getattr(message, "type", lambda: "console")() if callable(getattr(message, "type", None)) else getattr(message, "type", "console"),
                    "text": message.text,
                }
            ),
        )
        return page

    async def after_goto(page: Any, context: Any, url: str, response: Any, **_: Any) -> Any:
        del response
        try:
            state["final_url"] = page.url or url
            trace.log(f"Arrived at {state['final_url']}")
            await maybe_accept_cookie_banner(page, args.timeout_ms)
            await handle_investor_gate(page, args.role, args.timeout_ms, notes)
            await open_positionen_tab_if_needed(page, args.section_title, args.timeout_ms, notes)
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight * 0.45);")
            await page.wait_for_timeout(1200)
            await wait_for_holdings_section(page, args.section_title, args.timeout_ms)
            state["as_of_date"] = await extract_as_of_date(page, args.section_title)
            if not state["as_of_date"]:
                raise ScrapeError("Could not extract the holdings 'Stand:' date.")
            trace.log(f"Extracted holdings date: {state['as_of_date']}")

            state["cookies"] = await context.cookies()
            state["user_agent"] = await page.evaluate("navigator.userAgent")
            trace.log(f"Captured {len(state['cookies'])} cookies and user agent {state['user_agent']}")

            if trace.trace_dir:
                await save_page_snapshot(page, trace, "01_ready_before_export", include_html=True, include_text=True)
                page_summary = await collect_page_debug_summary(page, args.section_title, args.export_text)
                trace.save_json("01_ready_before_export_summary.json", page_summary)

            # ── Mark the export click boundary ──
            state["export_click_fired"] = True
            trace.log("Export click boundary set — subsequent network responses are post-click.")

            download = await trigger_export_download(page, args.section_title, args.export_text, args.timeout_ms, notes)

            # Wait for post-click network responses to arrive
            await page.wait_for_timeout(3000)

            try:
                state["export_audit"] = await page.evaluate("window.__invescoDownloadAudit || null")
            except Exception as exc:
                trace.log(f"Failed to read export audit object: {exc}")
            if trace.trace_dir:
                await save_page_snapshot(page, trace, "02_after_export_click", include_html=False, include_text=True)
                if state.get("export_audit") is not None:
                    trace.save_json("02_after_export_click_audit.json", state["export_audit"])
                page_summary = await collect_page_debug_summary(page, args.section_title, args.export_text)
                trace.save_json("02_after_export_click_summary.json", page_summary)

            if download is not None:
                try:
                    original_filename, saved_path = await save_playwright_download(download, output_dir, state["as_of_date"])
                    state["original_filename"] = original_filename
                    state["saved_path"] = saved_path
                    trace.remember(saved_path)
                    note("Captured a valid native browser download via Playwright.")
                    return page
                except Exception as exc:
                    note(f"Playwright download save failed validation: {exc}")

            blob_result = await try_save_blob_download(page, output_dir, state["as_of_date"])
            if blob_result is not None:
                original_filename, saved_path = blob_result
                state["original_filename"] = original_filename
                state["saved_path"] = saved_path
                trace.remember(saved_path)
                note("Recovered a valid XLSX from an in-page blob/object URL.")

            return page
        except Exception:
            if trace.trace_dir:
                await save_page_snapshot(page, trace, "99_after_goto_exception", include_html=True, include_text=True)
                try:
                    page_summary = await collect_page_debug_summary(page, args.section_title, args.export_text)
                    trace.save_json("99_after_goto_exception_summary.json", page_summary)
                except Exception as exc:
                    trace.log(f"Could not collect exception page summary: {exc}")
                try:
                    audit = await page.evaluate("window.__invescoDownloadAudit || null")
                    if audit is not None:
                        trace.save_json("99_after_goto_exception_audit.json", audit)
                except Exception as exc:
                    trace.log(f"Could not save exception audit object: {exc}")
            trace.save_text("traceback.txt", traceback.format_exc())
            raise

    crawler.crawler_strategy.set_hook("on_page_context_created", on_page_context_created)
    crawler.crawler_strategy.set_hook("after_goto", after_goto)

    try:
        await crawler.start()
        result = await crawler.arun(args.url, config=run_config)
    finally:
        await crawler.close()

    if state["response_tasks"]:
        await asyncio.gather(*state["response_tasks"], return_exceptions=True)

    if trace.trace_dir:
        trace.save_json("network_events.json", state["network_events"])
        if state["console_messages"]:
            trace.save_json("console_messages.json", state["console_messages"])
        html = getattr(result, "html", None)
        if isinstance(html, str) and html.strip():
            trace.save_text("crawler_result.html", html)

    candidate_urls = discover_candidate_urls(getattr(result, "network_requests", None))
    trace.log(f"Discovered {len(candidate_urls)} candidate URLs: {candidate_urls}")
    saved_path = state.get("saved_path")

    if not saved_path and getattr(result, "downloaded_files", None):
        for item in result.downloaded_files:
            if str(item).lower().endswith(".xlsx"):
                source = Path(item)
                body = source.read_bytes()
                if not is_probable_xlsx_bytes(body):
                    note(f"Ignored downloaded file that was not a valid XLSX workbook: {source.name}")
                    continue
                destination = output_dir / choose_saved_filename(state.get("as_of_date"), source.name)
                if source.resolve() != destination.resolve():
                    ensure_dir(destination.parent)
                    source.replace(destination)
                saved_path = str(destination.resolve())
                state["saved_path"] = saved_path
                state["original_filename"] = state.get("original_filename") or source.name
                trace.remember(saved_path)
                note("Recovered a valid XLSX from Crawl4AI result.downloaded_files.")
                break

    # ── Choose the best captured holdings payload ──
    best_captured = choose_best_captured_holdings(
        state["captured_holdings"],
        min_holdings=args.min_holdings,
    )

    if best_captured:
        state["holdings_payload"] = best_captured.payload
        state["holdings_url"] = best_captured.url
        state["holdings_summary"] = best_captured.summary
        note(
            f"Selected best holdings payload: "
            f"{'post-click' if best_captured.post_click else 'pre-click'}, "
            f"rows={best_captured.holdings_count}, "
            f"authoritative={payload_is_authoritative(best_captured.summary, args.min_holdings)}, "
            f"url={best_captured.url}"
        )
    else:
        # Log what was rejected and why
        for captured in state["captured_holdings"]:
            note(
                f"Rejected captured holdings payload as non-authoritative: "
                f"{'post-click' if captured.post_click else 'pre-click'}, "
                f"rows={captured.holdings_count}, "
                f"initial={'yes' if _is_initial_load_url(captured.url) else 'no'}, "
                f"url={captured.url}"
            )

    # ── Fallback A: derive full-data URLs from any captured initial-load URL ──
    if not state.get("holdings_payload"):
        initial_urls = [
            c.url for c in state["captured_holdings"] if _is_initial_load_url(c.url)
        ]
        derived_urls: list[str] = []
        for init_url in initial_urls:
            derived_urls.extend(_derive_full_holdings_urls(init_url))
        # Also include any non-initial candidate URLs discovered by the crawler
        holdings_candidate_url = choose_holdings_json_url(candidate_urls)
        if holdings_candidate_url and holdings_candidate_url not in derived_urls:
            derived_urls.append(holdings_candidate_url)

        for try_url in derived_urls:
            trace.log(f"Attempting full-holdings fetch from: {try_url}")
            try:
                fetched_payload = await fetch_json_from_url(
                    url=try_url,
                    cookies=state.get("cookies") or [],
                    user_agent=state.get("user_agent"),
                )
                fetched_summary = summarize_holdings_payload(fetched_payload)
                if payload_is_authoritative(fetched_summary, args.min_holdings):
                    state["holdings_payload"] = fetched_payload
                    state["holdings_url"] = try_url
                    state["holdings_summary"] = fetched_summary
                    note(
                        f"Fetched authoritative holdings JSON from API: "
                        f"rows={fetched_summary.get('holdings_count')}, url={try_url}"
                    )
                    break
                else:
                    note(
                        f"Rejected fetched holdings JSON as non-authoritative: "
                        f"rows={fetched_summary.get('holdings_count')}, "
                        f"declared_total={fetched_summary.get('declared_total')}, "
                        f"url={try_url}"
                    )
            except Exception as exc:
                note(f"JSON API fallback failed for {try_url}: {exc}")

    # ── Fallback B: accept the largest pre-click payload even if below min_holdings ──
    # This is a last-resort — it still builds the XLSX but marks it as partial.
    if not state.get("holdings_payload") and state["captured_holdings"]:
        largest = max(state["captured_holdings"], key=lambda c: c.holdings_count)
        if largest.holdings_count > 0:
            state["holdings_payload"] = largest.payload
            state["holdings_url"] = largest.url
            state["holdings_summary"] = largest.summary
            note(
                f"Accepted largest available holdings payload as PARTIAL fallback: "
                f"rows={largest.holdings_count}, "
                f"initial={'yes' if _is_initial_load_url(largest.url) else 'no'}, "
                f"url={largest.url}"
            )

    if trace.trace_dir and state.get("holdings_payload") is not None:
        trace.save_json("holdings_payload.json", state["holdings_payload"])
        if state.get("holdings_summary") is not None:
            trace.save_json("holdings_payload_summary.json", state["holdings_summary"])
    if trace.trace_dir and state["captured_holdings"]:
        trace.save_json(
            "all_captured_holdings_summary.json",
            [
                {
                    "url": c.url,
                    "post_click": c.post_click,
                    "holdings_count": c.holdings_count,
                    "is_initial": _is_initial_load_url(c.url),
                    "authoritative": payload_is_authoritative(c.summary, args.min_holdings),
                    "summary": c.summary,
                }
                for c in state["captured_holdings"]
            ],
        )

    if not saved_path and state.get("captured_binary"):
        binary = state["captured_binary"]
        if looks_like_xlsx_response(
            url=binary.get("url") or "",
            content_type=binary.get("content_type") or "",
            content_disposition=None,
            body=binary.get("body"),
        ):
            original_filename = binary.get("filename") or DEFAULT_ORIGINAL_FILENAME
            destination = output_dir / choose_saved_filename(state.get("as_of_date"), original_filename)
            destination.write_bytes(binary["body"])
            state["saved_path"] = str(destination.resolve())
            state["original_filename"] = original_filename
            saved_path = state["saved_path"]
            trace.remember(saved_path)
            note(f"Saved a validated XLSX captured from a network response: {binary.get('url')}")

    # ── Build XLSX from holdings payload ──
    if not saved_path and state.get("holdings_payload") and state.get("holdings_summary"):
        is_authoritative = payload_is_authoritative(state["holdings_summary"], args.min_holdings)
        holdings_count = state["holdings_summary"].get("holdings_count", 0)
        original_filename = state.get("original_filename") or DEFAULT_ORIGINAL_FILENAME
        destination = output_dir / choose_saved_filename(state.get("as_of_date"), original_filename)
        build_xlsx_from_holdings_payload(
            payload=state["holdings_payload"],
            source_url=args.url,
            holdings_url=state.get("holdings_url"),
            section_title=args.section_title,
            output_path=destination,
        )
        state["saved_path"] = str(destination.resolve())
        state["original_filename"] = original_filename
        saved_path = state["saved_path"]
        trace.remember(saved_path)
        if is_authoritative:
            note(
                "Constructed the XLSX locally from an authoritative holdings JSON payload "
                f"(rows={holdings_count})."
            )
        else:
            note(
                f"Constructed the XLSX locally from a PARTIAL holdings JSON payload "
                f"(rows={holdings_count}, min_required={args.min_holdings}). "
                "The payload did not pass completeness checks — the file may be incomplete."
            )

    if not saved_path:
        maybe_binary_url = next((u for u in candidate_urls if u.lower().endswith(".xlsx")), None)
        if maybe_binary_url:
            try:
                body, filename, content_type = await fallback_fetch_binary(
                    url=maybe_binary_url,
                    cookies=state.get("cookies") or [],
                    user_agent=state.get("user_agent"),
                )
                if body and is_probable_xlsx_bytes(body):
                    original_filename = filename or DEFAULT_ORIGINAL_FILENAME
                    destination = output_dir / choose_saved_filename(state.get("as_of_date"), original_filename)
                    destination.write_bytes(body)
                    state["saved_path"] = str(destination.resolve())
                    state["original_filename"] = original_filename
                    saved_path = state["saved_path"]
                    trace.remember(saved_path)
                    note(f"Downloaded a validated XLSX via binary URL fallback: {maybe_binary_url} ({content_type})")
            except Exception as exc:
                note(f"Binary URL fallback failed for {maybe_binary_url}: {exc}")

    # ── Freshness validation ──
    payload_date = normalize_date_ddmmyyyy(str((state.get("holdings_payload") or {}).get("effectiveDate") or ""))
    if payload_date and state.get("as_of_date") and payload_date != state.get("as_of_date"):
        note(f"Warning: page 'Stand:' date ({state['as_of_date']}) did not match holdings API effectiveDate ({payload_date}).")

    # Validate workbook date against page date if we saved a workbook
    state["workbook_summary"] = inspect_workbook(saved_path)
    if state.get("workbook_summary") and state.get("as_of_date"):
        wb_metadata = state["workbook_summary"].get("metadata") or {}
        wb_effective_date = normalize_date_ddmmyyyy(str(wb_metadata.get("effective_date") or ""))
        if wb_effective_date and wb_effective_date != state["as_of_date"]:
            note(
                f"Warning: workbook effective_date ({wb_effective_date}) does not match "
                f"page 'Stand:' date ({state['as_of_date']}). The workbook may be stale."
            )

    if trace.trace_dir and state.get("workbook_summary") is not None:
        trace.save_json("workbook_summary.json", state["workbook_summary"])

    completeness_ok, completeness_reasons = evaluate_completeness(
        workbook_summary=state.get("workbook_summary"),
        holdings_summary=state.get("holdings_summary"),
        min_holdings=args.min_holdings,
    )
    if completeness_ok is False:
        for reason in completeness_reasons:
            note(f"Warning: {reason}.")
        if trace.trace_dir:
            trace.log("Completeness check failed; trace artifacts kept for diagnosis.")
    elif completeness_ok is True:
        trace.log("Completeness check passed.")

    partial_success = bool(state.get("saved_path")) and bool(state.get("as_of_date") or payload_date)
    success = partial_success and completeness_ok is not False

    manifest = ScrapeManifest(
        source_url=args.url,
        final_url=state.get("final_url") or getattr(result, "url", None),
        as_of_date=state.get("as_of_date") or payload_date,
        role_selected=args.role,
        section_title=args.section_title,
        export_text=args.export_text,
        original_filename=state.get("original_filename"),
        saved_file=saved_path,
        xlsx_candidate_urls=candidate_urls,
        status_code=getattr(result, "status_code", None),
        success=success,
        partial_success=partial_success,
        completeness_ok=completeness_ok,
        expected_min_holdings=args.min_holdings,
        trace_dir=str(trace.trace_dir.resolve()) if trace.trace_dir else None,
        debug_files=trace.files,
        holdings_summary=state.get("holdings_summary"),
        workbook_summary=state.get("workbook_summary"),
        notes=notes,
    )

    manifest_path = output_dir / "manifest.json"
    manifest_path.write_text(json.dumps(asdict(manifest), ensure_ascii=False, indent=2), encoding="utf-8")
    trace.remember(manifest_path)
    trace.log(f"Wrote manifest to {manifest_path}")
    return manifest


# ── Self-test ───────────────────────────────────────────────────────────────

def run_self_test() -> None:
    # -- base scraper tests --
    sample_text = (
        "Positionen\n"
        "Die 10 größten Positionen\n"
        "Stand: 30.03.2026\n"
        "Daten exportieren\n"
    )
    assert extract_as_of_date_from_text(sample_text) == "30.03.2026"
    assert normalize_date_ddmmyyyy("2026-03-31") == "31.03.2026"
    assert sanitize_filename("Die_10_größten_Positionen-holdings.xlsx") == "Die_10_groessten_Positionen-holdings.xlsx"
    assert choose_saved_filename("30.03.2026", "Die_10_größten_Positionen-holdings.xlsx") == (
        "2026-03-30__Die_10_groessten_Positionen-holdings.xlsx"
    )

    fake_xlsx = io.BytesIO()
    with zipfile.ZipFile(fake_xlsx, "w") as archive:
        archive.writestr("[Content_Types].xml", "<Types />")
        archive.writestr("_rels/.rels", "<Relationships />")
        archive.writestr("xl/workbook.xml", "<workbook />")
    fake_xlsx_bytes = fake_xlsx.getvalue()

    assert is_probable_xlsx_bytes(fake_xlsx_bytes)
    assert not is_probable_xlsx_bytes(b"not-an-xlsx")
    assert looks_like_xlsx_response(
        url="https://example.com/file.bin",
        content_type="application/octet-stream",
        content_disposition='attachment; filename="holdings.xlsx"',
        body=fake_xlsx_bytes,
    )
    assert not looks_like_xlsx_response(
        url="https://example.com/image.jpeg",
        content_type="image/jpeg",
        content_disposition='attachment; filename="hero.jpeg"',
        body=b"\xff\xd8\xff",
    )

    events = [
        {
            "event_type": "response",
            "url": "https://example.com/files/Die_10_größten_Positionen-holdings.xlsx",
            "headers": {"content-type": XLSX_MIME},
        },
        {
            "event_type": "response",
            "url": "https://dng-api.invesco.com/cache/v1/accounts/de_DE/shareclasses/IE000716YHJ7/holdings/index?idType=isin&loadType=initial",
            "headers": {"content-type": JSON_MIME},
        },
        {
            "event_type": "response",
            "url": "https://example.com/static/hero.jpeg",
            "headers": {"content-type": "image/jpeg"},
        },
    ]
    urls = discover_candidate_urls(events)
    assert urls == [
        "https://example.com/files/Die_10_größten_Positionen-holdings.xlsx",
        "https://dng-api.invesco.com/cache/v1/accounts/de_DE/shareclasses/IE000716YHJ7/holdings/index?idType=isin&loadType=initial",
    ]

    # -- Updated test: choose_holdings_json_url now deprioritizes loadType=initial --
    # When only loadType=initial is available, it is still returned (last resort)
    assert choose_holdings_json_url(urls) == (
        "https://dng-api.invesco.com/cache/v1/accounts/de_DE/shareclasses/IE000716YHJ7/holdings/index?idType=isin&loadType=initial"
    )
    # When a non-initial URL is also available, it should be preferred
    urls_with_export = urls + [
        "https://dng-api.invesco.com/cache/v1/accounts/de_DE/shareclasses/IE000716YHJ7/holdings/index?idType=isin&loadType=full",
    ]
    chosen = choose_holdings_json_url(urls_with_export)
    assert chosen == (
        "https://dng-api.invesco.com/cache/v1/accounts/de_DE/shareclasses/IE000716YHJ7/holdings/index?idType=isin&loadType=full"
    ), f"Expected non-initial URL to be preferred, got: {chosen}"

    # -- looks_like_holdings_json_url now matches broader /holdings/ pattern --
    assert looks_like_holdings_json_url(
        "https://dng-api.invesco.com/cache/v1/accounts/de_DE/shareclasses/IE000716YHJ7/holdings/index?idType=isin&loadType=initial"
    )
    assert looks_like_holdings_json_url(
        "https://dng-api.invesco.com/cache/v1/accounts/de_DE/shareclasses/IE000716YHJ7/holdings/index?idType=isin&loadType=full"
    )
    assert looks_like_holdings_json_url(
        "https://dng-api.invesco.com/cache/v1/accounts/de_DE/shareclasses/IE000716YHJ7/holdings/export"
    )
    assert not looks_like_holdings_json_url("https://example.com/static/hero.jpeg")

    # -- _is_initial_load_url helper --
    assert _is_initial_load_url(
        "https://dng-api.invesco.com/cache/v1/accounts/de_DE/shareclasses/IE000716YHJ7/holdings/index?idType=isin&loadType=initial"
    )
    assert not _is_initial_load_url(
        "https://dng-api.invesco.com/cache/v1/accounts/de_DE/shareclasses/IE000716YHJ7/holdings/index?idType=isin&loadType=full"
    )

    # -- _derive_full_holdings_urls helper --
    initial_url = "https://dng-api.invesco.com/cache/v1/accounts/de_DE/shareclasses/IE000716YHJ7/holdings/index?idType=isin&loadType=initial"
    derived = _derive_full_holdings_urls(initial_url)
    assert len(derived) >= 1, f"Expected at least 1 derived URL, got {derived}"
    assert any("loadType=full" in u for u in derived), f"Expected a loadType=full URL in {derived}"
    assert all("loadType=initial" not in u for u in derived), f"Derived URLs should not contain loadType=initial: {derived}"
    # Non-initial URL should produce no derivations
    assert _derive_full_holdings_urls("https://example.com/holdings/index?idType=isin&loadType=full") == []
    print("_derive_full_holdings_urls tests passed.")

    payload = {
        "effectiveDate": "2026-03-31",
        "holdings": [
            {"name": "NVIDIA CORP USD0.001", "cusip": "67066G104", "isin": "US67066G1040", "weight": 4.4628},
            {"name": "APPLE INC USD0.00001", "cusip": "037833100", "isin": "US0378331005", "weight": 3.9891},
        ],
    }
    temp_dir = Path("/tmp/invesco_holdings_selftest")
    ensure_dir(temp_dir)
    output_path = temp_dir / "holdings.xlsx"
    build_xlsx_from_holdings_payload(payload, DEFAULT_URL, "https://example.com/holdings/index", DEFAULT_SECTION_TITLE, output_path)
    assert output_path.exists() and output_path.stat().st_size > 0
    assert is_probable_xlsx_bytes(output_path.read_bytes())
    print("Base self-test passed.")

    # -- traced extension tests --
    payload_ext = {
        "effectiveDate": "2026-03-31",
        "totalCount": 2601,
        "holdings": [
            {"name": "NVIDIA", "cusip": "67066G104", "isin": "US67066G1040", "weight": 4.4628},
            {"name": "APPLE", "cusip": "037833100", "isin": "US0378331005", "weight": 3.9891},
        ],
    }
    payload_summary = summarize_holdings_payload(payload_ext)
    assert payload_summary["holdings_count"] == 2
    assert payload_summary["declared_total"] == 2601
    assert payload_summary["declared_total_exceeds_rows"] is True

    temp_dir2 = Path("/tmp/invesco_holdings_traced_selftest")
    ensure_dir(temp_dir2)
    workbook_path = temp_dir2 / "holdings.xlsx"
    build_xlsx_from_holdings_payload(payload_ext, DEFAULT_URL, "https://example.com/holdings/index", DEFAULT_SECTION_TITLE, workbook_path)
    workbook_summary = inspect_workbook(workbook_path)
    assert workbook_summary is not None
    assert workbook_summary["data_row_count"] == 2

    completeness_ok, reasons = evaluate_completeness(
        workbook_summary=workbook_summary,
        holdings_summary=payload_summary,
        min_holdings=2500,
    )
    assert completeness_ok is False
    assert reasons

    # -- payload_is_authoritative tests --
    assert not payload_is_authoritative(payload_summary, 2500), \
        "Payload with 2 rows should not be authoritative for min_holdings=2500"
    assert payload_is_authoritative({"holdings_count": 2600, "declared_total": 2600}, 2500)
    assert not payload_is_authoritative({"holdings_count": 2600, "declared_total": 3000}, 2500), \
        "Payload with declared_total > holdings_count should not be authoritative"
    assert payload_is_authoritative({"holdings_count": 2600, "declared_total": None}, 2500)
    assert payload_is_authoritative({"holdings_count": 2600}, 2500)
    assert not payload_is_authoritative({"holdings_count": 100}, 2500)
    print("payload_is_authoritative tests passed.")

    # -- choose_best_captured_holdings tests --
    summary_small = {"holdings_count": 100, "declared_total": None}
    summary_big = {"holdings_count": 2600, "declared_total": None}
    summary_partial = {"holdings_count": 2600, "declared_total": 4000}

    # Post-click authoritative wins over pre-click authoritative
    c_pre_auth = CapturedHoldings(payload={}, url="pre", summary=summary_big, post_click=False, holdings_count=2600)
    c_post_auth = CapturedHoldings(payload={}, url="post", summary=summary_big, post_click=True, holdings_count=2600)
    assert choose_best_captured_holdings([c_pre_auth, c_post_auth], 2500) is c_post_auth

    # Post-click non-authoritative wins over pre-click non-authoritative
    c_pre_small = CapturedHoldings(payload={}, url="pre", summary=summary_small, post_click=False, holdings_count=100)
    c_post_small = CapturedHoldings(payload={}, url="post", summary=summary_small, post_click=True, holdings_count=100)
    best = choose_best_captured_holdings([c_pre_small, c_post_small], 2500)
    assert best is c_post_small

    # Pre-click non-authoritative alone returns None (the root cause fix!)
    assert choose_best_captured_holdings([c_pre_small], 2500) is None

    # Pre-click authoritative is accepted when no post-click exists
    assert choose_best_captured_holdings([c_pre_auth], 2500) is c_pre_auth

    # Partial payload (declared_total > rows) is not authoritative
    c_partial = CapturedHoldings(payload={}, url="partial", summary=summary_partial, post_click=False, holdings_count=2600)
    assert choose_best_captured_holdings([c_partial], 2500) is None

    print("choose_best_captured_holdings tests passed.")

    # -- output-dir date sub-directory test --
    test_root = Path("/tmp/invesco_output_dir_test")
    if test_root.exists():
        shutil.rmtree(test_root)
    run_dir = resolve_run_output_dir(test_root)
    today = datetime.now().strftime("%Y_%m_%d")
    assert run_dir == (test_root.resolve() / today)
    assert run_dir.is_dir()
    # running again replaces cleanly
    (run_dir / "dummy.txt").write_text("hello")
    run_dir2 = resolve_run_output_dir(test_root)
    assert run_dir2 == run_dir
    assert not (run_dir2 / "dummy.txt").exists()
    shutil.rmtree(test_root)

    print("Extended trace self-test passed.")
    print("All self-tests passed.")


# ── Entry points ────────────────────────────────────────────────────────────

async def async_main() -> int:
    args = parse_args()
    if args.self_test:
        run_self_test()
        return 0
    manifest = await run_scrape(args)
    print(json.dumps(asdict(manifest), ensure_ascii=False, indent=2))
    return 0 if manifest.success else 1


def main() -> None:
    try:
        raise SystemExit(asyncio.run(async_main()))
    except KeyboardInterrupt:
        raise SystemExit(130)


if __name__ == "__main__":
    main()